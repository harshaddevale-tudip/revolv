import csv
from collections import OrderedDict
import logging

from django.contrib import messages
from django.contrib.auth.models import User
from django.conf import settings
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout as auth_logout
from django.contrib.auth import views as auth_views
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.core.urlresolvers import reverse
from django.db.models import Sum
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.http.response import HttpResponseBadRequest, HttpResponseRedirect
from django.shortcuts import redirect, render_to_response, get_object_or_404
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.debug import sensitive_post_parameters
from django.views.decorators.http import require_http_methods
from django.views.generic import FormView, TemplateView, View
from django.template.context import RequestContext
from revolv.base.forms import SignupForm, AuthenticationForm
from revolv.base.users import UserDataMixin
from revolv.base.utils import ProjectGroup
from revolv.payments.models import Payment, Tip, PaymentType, RepaymentFragment
from revolv.project.models import Category, Project, ProjectMatchingDonors
from revolv.project.utils import aggregate_stats
from revolv.donor.views import humanize_integers, total_donations
from revolv.base.models import RevolvUserProfile
from revolv.tasks.sfdc import send_signup_info
from revolv.lib.mailer import send_revolv_email
from social.apps.django_app.default.models import UserSocialAuth
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate
from django.core import serializers
import mailchimp
import json
import re

logger = logging.getLogger(__name__)
LIST_ID = settings.LIST_ID

class HomePageView(UserDataMixin, TemplateView):
    """
    Website home page.

    TODO: this view is deprecated - most of the context variables are not
    used anymore. Should be cleaned up.
    """
    template_name = 'base/home.html'
    FEATURED_PROJECT_TO_SHOW = 6

    def get_global_impacts(self):
        """
        Returns: A dictionary of RE-volv wide impact figures.
        """
        #carbon_saved_by_month = Project.objects.statistics().pounds_carbon_saved_per_month
        # Assume 20 year lifetime.
        # We use str() to avoid django adding commas to integer in the template.
        #carbon_saved = str(int(carbon_saved_by_month * 12 * 20))
        people_donated_sys_count = RevolvUserProfile.objects.exclude(project=None).count()
        people_donated_stat_Count = str(int(people_donated_sys_count + 615))
        #total_kwh = float(Project.objects.aggregate(n=Sum('total_kwh_value'))['n'])
        #carbon_value_calc = total_kwh * 1.5
        #funding_goal_value = float(Project.objects.aggregate(n=Sum('funding_goal'))['n'])
	#amount_invested_value = Payment.objects.aggregate(n=Sum('amount'))['n']
        #final_carbon_avoided = str(int(carbon_value_calc / funding_goal_value * amount_invested_value + 3057319))
        global_impacts = {
            # Users who have backed at least one project:
            'num_people_donated': people_donated_stat_Count,
            'num_projects': len(Project.objects.get_completed()),
            'num_people_affected':
                Project.objects.filter(project_status=Project.COMPLETED).aggregate(n=Sum('people_affected'))['n'],
            'co2_avoided': str(int(Project.objects.get_total_avoided_co2())),
            #	'co2_avoided': 7452670,
        }
        return global_impacts

    def get_context_data(self, **kwargs):
        context = super(HomePageView, self).get_context_data(**kwargs)
        featured_projects = Project.objects.get_featured(HomePageView.FEATURED_PROJECT_TO_SHOW)
        active_projects = Project.objects.get_active()
        context["active_projects"] = filter(lambda p: p.amount_left > 0.0, active_projects)
        completed_projects = Project.objects.get_completed()
        context["first_project"] = active_projects[0] if len(active_projects) > 0 else None
        # Get top 6 featured projects, Changed to active Projects in final fix
        context["featured_projects"] = active_projects[:6]
        #accept return value from project/model.py and display it on project/home.html file
        context["completed_featured_projects"] = completed_projects[:6]
        context["completed_projects_count"] = Project.objects.get_completed().count()
        context["total_donors_count"] = Payment.objects.total_distinct_organic_donors()
        context["global_impacts"] = self.get_global_impacts()
        return context

class DonationReportView(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    model = Payment
    template_name = 'base/partials/donation_report.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if not request.user.revolvuserprofile.is_administrator():
            return HttpResponseRedirect(reverse("dashboard"))
        return super(DonationReportView, self).dispatch(request, *args, **kwargs)

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(DonationReportView, self).get_context_data(**kwargs)
        return context


class MatchingDonorsView(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    model = Payment
    template_name = 'base/partials/matching_donors.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if not request.user.revolvuserprofile.is_administrator():
            return HttpResponseRedirect(reverse("dashboard"))
        return super(MatchingDonorsView, self).dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # self.object = self.get_object()
        context = super(MatchingDonorsView, self).get_context_data(**kwargs)
        context['donors'] = ProjectMatchingDonors.objects.all()
        context['users'] = RevolvUserProfile.objects.all()
        context['projects'] = Project.objects.all()
        return self.render_to_response(context)

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(MatchingDonorsView, self).get_context_data(**kwargs)
        context['donors'] = ProjectMatchingDonors.objects.all()
        context['users'] = RevolvUserProfile.objects.all()
        context['projects'] = Project.objects.all()
        return context


class DonationReportForProject(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    model = Payment
    template_name = 'base/partials/ambassador_donation_report.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if not request.user.revolvuserprofile.is_ambassador():
            return HttpResponseRedirect(reverse("dashboard"))
        return super(DonationReportForProject, self).dispatch(request, *args, **kwargs)
    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(DonationReportForProject, self).get_context_data(**kwargs)
        return context

class RepaymentReport(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    model = Payment
    template_name = 'base/partials/repayment_report.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if not request.user.revolvuserprofile.is_administrator():
            return HttpResponseRedirect(reverse("dashboard"))
        return super(RepaymentReport, self).dispatch(request, *args, **kwargs)
    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(RepaymentReport, self).get_context_data(**kwargs)
        return context

class BaseStaffDashboardView(UserDataMixin, TemplateView):
    """
    Base view for the administrator and ambassador dashboard views. The
    specific views in administrator/views.py and ambassador/views.py
    will inherit from this view.
    """

    def get_filter_args(self):
        """
        Return an array of arguments to pass to Project.objects.get_[drafted|proposed|active|completed].
        """
        return []

    def get_context_data(self, **kwargs):
        context = super(BaseStaffDashboardView, self).get_context_data(**kwargs)

        project_dict = OrderedDict()
        project_dict[ProjectGroup('Proposed Projects', "proposed")] = Project.objects.get_proposed(*self.get_filter_args())
        project_dict[ProjectGroup('Staged projects', "staged")] = Project.objects.get_staged(*self.get_filter_args())
        project_dict[ProjectGroup('Active Projects', "active")] = Project.objects.get_active(*self.get_filter_args())
        project_dict[ProjectGroup('Completed Projects', "completed")] = Project.objects.get_completed(*self.get_filter_args())

        context["project_dict"] = project_dict
        context["role"] = self.role or "donor"

        context['donated_projects'] = Project.objects.donated_projects(self.user_profile)
        statistics_dictionary = aggregate_stats(self.user_profile)
        statistics_dictionary['total_donated'] = total_donations(self.user_profile)
        total_people_affected = Project.objects.donated_completed_projects(self.user_profile)
        statistics_dictionary['people_served'] = total_people_affected
        humanize_integers(statistics_dictionary)
        context['statistics'] = statistics_dictionary

        return context


class CategoryPreferenceSetterView(UserDataMixin, View):
    http_methods = ['post']

    def http_method_not_allowed(request, *args, **kwargs):
        return redirect("dashboard")

    def post(self, request, *args, **kwargs):
        user = self.user_profile
        user.preferred_categories.clear()
        info_dict = request.POST.dict()
        for category_string in info_dict:
            category = Category.objects.get(id=category_string)
            user.preferred_categories.add(category)
        return HttpResponse()


class ProjectListView(UserDataMixin, TemplateView):
    """ Base View of all active projects
    """
    model = Project
    template_name = 'base/projects-list.html'

    def get_context_data(self, **kwargs):
        context = super(ProjectListView, self).get_context_data(**kwargs)
        active = Project.objects.get_active()
        context["active_projects"] = filter(lambda p: p.amount_left > 0.0, active)
        # context["active_projects"] = active
        context["is_reinvestment"] = False
        return context



class SignInView(TemplateView):
    """Signup and login page. Has three submittable forms: login, signup,
    and sign in with facebook.

    Note that the signup with facebook functionality will automatically
    create a user object, but if the user has previously signed up with
    facebook, they will not be able to sign in without facebook to the
    same account.

    Also note that the "sign in with facebook" does not necessarily do
    the same thing as either sign up or login: when the user clicks this
    button, they will be automatically signed up and logged in if there
    is not an account associated with their facebook profile, or they will
    just be logged in if there is.

    Http verbs:
        GET: renders the signin page with empty forms.
    """
    template_name = "base/sign_in.html"
    login_form_class = AuthenticationForm
    signup_form_class = SignupForm

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated():
            return redirect("home")
        return super(SignInView, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, *args, **kwargs):
        context = super(SignInView, self).get_context_data(**kwargs)
        login_form = self.login_form_class()
        signup_form = self.signup_form_class()
        context["signup_form"] = signup_form
        context["login_form"] = login_form
        if self.request.GET.get("next"):
            context["login_redirect_url"] = self.request.GET.get("next")
        else:
            context["login_redirect_url"] = reverse('dashboard')
        context["referring_endpoint"] = ""
        context["reason"] = self.request.GET.get("reason")
        return context


class RedirectToSigninOrHomeMixin(object):
    """
    Mixin that detects if a page was requested with method="GET", and redirects
    to the signin page if so. Also, if posted to with an already authenticated
    user, will redirect to the homepage instead.

    This is useful both for the login and signup endpoints.
    """
    @method_decorator(sensitive_post_parameters(
        "password", "password1", "password2"
    ))
    def dispatch(self, request, *args, **kwargs):
        # don't allow rendering a form page with a GET request to this view:
        # instead, redirect to the signin page
        if self.request.method == "GET":
            return redirect("signin")
        # if the user is already logged in and tries to log in again, just
        # redirect them to the home page.
        if request.user.is_authenticated():
            return redirect("home")
        return super(RedirectToSigninOrHomeMixin, self).dispatch(
            request, *args, **kwargs
        )


class LoginView(RedirectToSigninOrHomeMixin, FormView):
    """
    Login endpoint: checks the data from the received request against
    django.contrib.auth.forms.AuthenticationForm and logs in the user if
    possible. If not, redirects back to the signin page.

    Http verbs:
        GET: redirect to signin page
        POST: check post parameters for user credentials, login the user
            and redirect to the specified next page (home by default), or
            render the sign in page with errors.
    """
    form_class = AuthenticationForm
    template_name = 'base/sign_in.html'
    url_append = "#login"
    redirect_view = "signin"

    @csrf_exempt
    @method_decorator(sensitive_post_parameters('password'))
    def dispatch(self, request, *args, **kwargs):
        self.next_url = request.POST.get("next", "home")
        if request.POST.get("payment"):
            self.request.session['payment'] = request.POST.get("payment")
        return super(LoginView, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        """Log the user in and redirect them to the supplied next page."""
        self.username = self.request.POST.get('username')
        self.password = self.request.POST.get('password')
        """Log the user in and redirect them to the supplied next page."""
        user = authenticate(username=self.username, password=self.password)

        auth_login(self.request, form.get_user())
        if self.request.session.get('payment'):
            payment_id=self.request.session.get('payment')
            # amount = self.amount
            # tip = self.tip
            user=RevolvUserProfile.objects.get(user=self.request.user)
            if user:
                Payment.objects.filter(id=payment_id).update(user=user, entrant=user)
                payment=Payment.objects.get(pk=payment_id)
                project=Project.objects.get(id=payment.project_id)
                project.donors.add(user)
                guest_user = User.objects.get(username='Guest')
                guest_profile = RevolvUserProfile.objects.get(user=guest_user)
                payment_count = Payment.objects.filter(user=guest_profile,project=project).count()
                if payment_count <= 1:
                    project.donors.remove(guest_profile)
                Tip.objects.filter(id=payment.tip_id).update(user=user)
                del self.request.session['payment']
                messages.success(self.request, 'Logged in as ' + self.request.POST.get('username'))
                return redirect(reverse('dashboard'))
        messages.success(self.request, 'Logged in as ' + self.request.POST.get('username'))
        return redirect(self.next_url)

    def get_context_data(self, *args, **kwargs):
        context = super(LoginView, self).get_context_data(*args, **kwargs)
        context["signup_form"] = SignupForm()
        context["login_form"] = self.get_form(self.form_class)
        context["referring_endpoint"] = "login"
        return context


class SignupView(RedirectToSigninOrHomeMixin, FormView):
    """
    Signup endpoint: processes the signup form and signs the user up (and logs
    them in). Note that the sign up with facebook functionality is entirely
    different: this is only for the regular django auth signup flow.

    Http verbs:
        GET: redirect to signin page
        POST: check post params against form, redirect to signin page if the
            form is not valid.
    """
    form_class = SignupForm
    template_name = "base/sign_in.html"
    url_append = "#signup"
    redirect_view = "signin"

    @csrf_exempt
    def dispatch(self, request, *args, **kwargs):
        self.next_url = request.POST.get("next", "home")
        if request.user.is_authenticated():
            return redirect("dashboard")
        if request.POST.get("payment"):
            self.request.session['payment'] = request.POST.get("payment")
        return super(SignupView, self).dispatch(request, *args, **kwargs)

    @csrf_exempt
    def form_valid(self, form):
        form.save()
        u = form.ensure_authenticated_user()
        name = u.revolvuserprofile.get_full_name()
        send_signup_info(name, u.email, u.revolvuserprofile.address)

        # log in the newly created user model. if there is a problem, error
        auth_login(self.request, u)
        SITE_URL = settings.SITE_URL
        login_link = SITE_URL + reverse('login')
        portfolio_link = SITE_URL + reverse('dashboard')
        context = {}
        context['user'] = self.request.user
        context['login_link'] = login_link
        context['portfolio_link'] = portfolio_link
        # send_revolv_email(
        #     'signup',
        #     context, [self.request.user.email]
        # )

        user = RevolvUserProfile.objects.get(user=self.request.user)
        if user.subscribed_to_newsletter:
            try:
                is_email_exist = False
                list = mailchimp.utils.get_connection().get_list_by_id(LIST_ID)
                for resp in list.con.list_members(list.id)['data']:
                    if self.request.user.email == resp['email']:
                        is_email_exist = True
                if is_email_exist:
                    pass
                else:
                    list.subscribe(self.request.user.email, {'EMAIL': self.request.user.email})

            except Exception, e:
                logger.exception(e)



        if self.request.session.get('payment'):
            payment_id=self.request.session.get('payment')
            user=RevolvUserProfile.objects.get(user=self.request.user)
            if user:
                Payment.objects.filter(id=payment_id).update(user=user, entrant=user)
                payment = Payment.objects.get(pk=payment_id)
                project = Project.objects.get(id=payment.project_id)
                project.donors.add(user)
                guest_user = User.objects.get(username='Guest')
                guest_profile = RevolvUserProfile.objects.get(user=guest_user)
                payment_count = Payment.objects.filter(user=guest_profile, project=project).count()
                if payment_count <= 1:
                    project.donors.remove(guest_profile)
                Tip.objects.filter(id=payment.tip_id).update(user=user)
                del self.request.session['payment']
                messages.success(self.request, 'Signed up as ' + self.request.POST.get('username'))
                return redirect(reverse('dashboard'))
        messages.success(self.request, 'Signed up successfully!')
        # return redirect("dashboard" +'?social=true')
        return HttpResponseRedirect(reverse("dashboard") + '?social=signup')

    def get_context_data(self, *args, **kwargs):
        context = super(SignupView, self).get_context_data(**kwargs)
        context["signup_form"] = self.get_form(self.form_class)
        context["login_form"] = AuthenticationForm()
        context["referring_endpoint"] = "signup"
        return context


class LogoutView(UserDataMixin, View):
    """
    Basic logout view: Accessed whenever the user wants to logout, processes
    the logout, shows a toast, and redirects to home.
    """

    def get(self, request, *args, **kwargs):
        auth_logout(request)
        messages.success(self.request, 'Logged out successfully')
        return redirect('home')

class ReinvestmentRedirect(UserDataMixin, TemplateView):

    model = Project
    template_name = 'base/projects-list.html'

    def get_object(self):
        return self.model.objects.get(pk=self.request.user.pk)

    def get_context_data(self, **kwargs):
        context = super(ReinvestmentRedirect, self).get_context_data(**kwargs)
        active = Project.objects.get_active()
        context["active_projects"] = filter(lambda p: p.amount_left > 0.0, active)
        if self.user_profile and self.user_profile.reinvest_pool > 0.0:
            context["is_reinvestment"] = True
            context["reinvestment_amount"] = self.user_profile.reinvest_pool

        return context

    # def get(self, request, *args, **kwargs):
    #     if self.is_administrator:
    #         return render_to_response('base/partials/project.html',context_instance=RequestContext(request))

def solarathome(request):
    return render_to_response('base/solar_at_home.html',
                              context_instance=RequestContext(request))

def bring_solar_tou_your_community(request):
    return render_to_response('base/bring_solar_at_community.html',
                              context_instance=RequestContext(request))

def intake_form_submit(request):
    try:
        # projectData = request.POST['projectData']
        email = request.POST.get('email')

    except:
        logger.exception('Form values are not valid')
        return HttpResponseBadRequest('bad POST data')


def select_chapter(request, chapter):
    if chapter == '1':
        return render_to_response('base/chapter.html',
                              context_instance=RequestContext(request))
    if chapter == '2':
        return render_to_response('base/chapter2.html',
                                  context_instance=RequestContext(request))
    if chapter == '3':
        return render_to_response('base/chapter3.html',
                                  context_instance=RequestContext(request))
    if chapter == '4':
        return render_to_response('base/chapter4.html',
                                  context_instance=RequestContext(request))
    if chapter == '5':
        return render_to_response('base/chapter5.html',
                                  context_instance=RequestContext(request))
    if chapter == '6':
        return render_to_response('base/chapter6.html',
                                  context_instance=RequestContext(request))
    if chapter == '7':
        return render_to_response('base/chapter7.html',
                                  context_instance=RequestContext(request))
    if chapter == '8':
        return render_to_response('base/chapter8.html',
                                  context_instance=RequestContext(request))
    if chapter == '9':
        return render_to_response('base/chapter9.html',
                                  context_instance=RequestContext(request))
    if chapter == '10':
        return render_to_response('base/chapter10.html',
                                  context_instance=RequestContext(request))
    if chapter == '11':
        return render_to_response('base/chapter11.html',
                                  context_instance=RequestContext(request))
    if chapter == '12':
        return render_to_response('base/chapter1'
                                  '2.html',
                                  context_instance=RequestContext(request))

def intake_form(request):
    return render_to_response('base/intake_form.html',
                              context_instance=RequestContext(request))

class DashboardRedirect(UserDataMixin, View):
    """
    Redirects user to appropriate dashboard. (e.g. Administrators automagically
    go to the /my-portfolio/admin endpoint)

    Redirects to home page if not authenticated.
    """

    def get(self, request, *args, **kwargs):
        social = request.GET.get('social', '')
        amount = request.session.get('amount')
        project = request.session.get('project')
        cover_photo = request.session.get('cover_photo','')
        if bool(request.GET) is False :
            if not self.is_authenticated:
                return redirect('home')
            if self.is_administrator:
                return redirect('administrator:dashboard')
            if self.is_ambassador:
                return redirect('ambassador:dashboard')
            return redirect(reverse('donor:dashboard'))

        elif social == 'donation':
            if not self.is_authenticated:
                return redirect('home')

            if self.is_administrator:
                return redirect('administrator:dashboard')

            if self.is_ambassador:
                return redirect(reverse('ambassador:dashboard') + '?social=' + social+'&amount='+amount+'&project='+project+'&cover_photo='+''.join(cover_photo))
            if self.is_donor:
                return redirect(reverse('donor:dashboard') + '?social=' + social+'&amount='+amount+'&project='+project+'&cover_photo='+''.join(cover_photo))

        else:
            if not self.is_authenticated:
                return redirect('home')

            if self.is_administrator:
                return redirect('administrator:dashboard')

            if self.is_ambassador:
                return redirect(reverse('ambassador:dashboard') + '?social=' + social)
            if self.is_donor:
                return redirect(reverse('donor:dashboard') + '?social=' + social)


# password reset/change views: thin wrappers around django's built in password
# reset views, but with our own templates
def password_reset_initial(request):
    return auth_views.password_reset(
        request,
        template_name="base/auth/forgot_password_initial.html",
        email_template_name="base/auth/forgot_password_email.html",
        from_email="support@re-volv.org"
    )


def password_change(request):
    return auth_views.password_change(
        request,
        template_name="base/auth/change_password.html",
        post_change_redirect="/my-portfolio/donor/?password_change_success",
    )


def password_reset_done(request):
    return auth_views.password_reset_done(request, template_name="base/auth/forgot_password_done.html")


def password_reset_confirm(request, *args, **kwargs):
    kwargs.update({"template_name": "base/auth/forgot_password_confirm.html"})
    return auth_views.password_reset_confirm(request, *args, **kwargs)


def password_reset_complete(request):
    return auth_views.password_reset_complete(request, template_name="base/auth/forgot_password_complete.html")


@login_required
@require_http_methods(['GET'])
def unsubscribe(request, action):
    """
    View handle unsubscribe email update
    """
    data = {}
    if action and action.lower() == 'updates':
        user_profile = request.user.revolvuserprofile
        user_profile.subscribed_to_updates = False
        user_profile.save()
        data = {'msg': "You have successfully unsubscribed"}
    else:
        data = {'msg': 'Please specify which section you want to unsubscribe'}

    return render_to_response('base/minimal_message.html',
                              context_instance=RequestContext(request, data))

@login_required
def delete(request):
    pk=request.REQUEST['id']
    ProjectMatchingDonor = get_object_or_404(ProjectMatchingDonors, pk=pk)
    ProjectMatchingDonor.delete()
    return HttpResponse(json.dumps({'status': 'delete'}), content_type="application/json")

@login_required
def edit(request):
    pk=request.REQUEST['id']
    ProjectMatchingDonor = get_object_or_404(ProjectMatchingDonors, pk=pk)
    serialized_obj = serializers.serialize('json', [ProjectMatchingDonor, ])
    return HttpResponse(json.dumps({'ProjectMatchingDonor': serialized_obj}), content_type="application/json")

@login_required
def social_connection(request):
    """
    View handle my social connection page
    """
    backend_map = {'facebook': {'name': 'facebook', 'connected': False,
                                'dc_url': reverse('social:disconnect', kwargs={'backend': 'facebook'})},
                   'google': {'name': 'google-oauth2', 'connected': False,
                              'dc_url': reverse('social:disconnect', kwargs={'backend': 'google-oauth2'})}
                   }
    accounts = UserSocialAuth.objects.filter(user=request.user)

    for account in accounts:
        for k, v in backend_map.iteritems():
                if v['name'] == account.provider:
                    backend_map[k]['connected'] = True

    return render_to_response('base/social_account.html',
                              context_instance=RequestContext(request, {'accounts': backend_map}))

@login_required
@require_http_methods(['POST'])
def add_maching_donors(request):
    id=request.POST.get('id')
    user=request.POST['user']
    project=request.POST['project']
    amount=request.POST['amount']
    revolv_user=get_object_or_404(RevolvUserProfile, pk=user)

    matching_project=get_object_or_404(Project, pk=project)

    if not id:
        ProjectMatchingDonors.objects.create(
            matching_donor=revolv_user,
            project=matching_project,
            amount=float(amount)
        )
    else:
        projectMatchingDonor=ProjectMatchingDonors.objects.get(id=id)
        projectMatchingDonor.matching_donor=revolv_user
        projectMatchingDonor.project=matching_project
        projectMatchingDonor.amount=amount
        projectMatchingDonor.save()

    return HttpResponse(json.dumps({'status': 'created'}), content_type="application/json")



def social_exception(request):
    has_social_exception = request.session.get('has_social_exception')
    if not has_social_exception:
        return redirect(reverse('home'))

    del request.session['has_social_exception']

    message = request.GET.get('message')
    return render_to_response('base/minimal_message.html',
                              context_instance=RequestContext(request, {'msg': message}))


def matching_donor_reinvestment(request):
    pk=request.GET.get('id')
    with open('/home/tudip/Desktop/Admin_reinvestment_on_15th.csv') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            amount=row[9]
            project_name=re.sub('-AC$', '', row[4])
            #project = get_object_or_404(Project, pk=pk)
            project=Project.objects.get(pk=pk)
            project_matching_donors = ProjectMatchingDonors.objects.filter(project=project, amount__gt=0)
            if project.title == project_name:
                if project_matching_donors:
                    donor=ProjectMatchingDonors.objects.get(project=project, amount__gt=0)

                    try:
                        if donor.amount > float(amount):
                            matching_donation = float(amount)
                            donor.amount = donor.amount - float(amount)
                            donor.save()
                        else:
                            matching_donation = donor.amount
                            donor.amount = 0
                            donor.save()

                        tip = None

                        project_assign = Project.objects.get(title=project_name)

                        Payment.objects.create(
                            user=donor.matching_donor,
                            entrant=donor.matching_donor,
                            amount=matching_donation,
                            project=project_assign,
                            tip=tip,
                            payment_type=PaymentType.objects.get_stripe(),
                        )

                    except Exception as e:
                        print "~~~~~~~~~ exception ~~~~~~~~", e

    return HttpResponseRedirect(reverse("home"))


def ambassador_data_table(request):
    draw=request.GET.get('draw')
    datepicker1=request.GET.get('datepicker1')
    datepicker2=request.GET.get('datepicker2')
    length=request.GET.get('length')
    order=request.GET.get('order[0][dir]')
    start=request.GET.get('start')
    search=request.GET.get('search[value]')
    currentSortByCol = request.GET.get('order[0][column]')


    fields=['user__user__first_name', 'user__user__last_name', 'user__user__username','user__user__email','created_at','project__title','amount','user_reinvestment__amount','admin_reinvestment__amount','tip__amount']

    order_by={'desc':'-','asc':''}

    column_order=order_by[order]+fields[int(currentSortByCol)]

    revolv_user = get_object_or_404(RevolvUserProfile, pk=request.user.id)

    projects = Project.objects.all()
    project_list=[]
    for project in projects:
        for ambassador in project.ambassadors.all():
            if revolv_user == ambassador:
                project_list.append(project)


    payment_list = Payment.objects.filter(project__in=project_list).order_by((column_order))


    if search.strip():
        payment_list = payment_list.filter(Q(user__user__username__icontains=search)|
                                              Q(user__user__first_name__icontains=search)|
                                              Q(project__title__icontains=search)|
                                              Q(amount__icontains=search)|
                                              Q(user__user__last_name__icontains=search)|
                                              Q(user__user__email__icontains=search))


    if datepicker1 and datepicker2:
        import datetime
        import pytz

        date1 = datetime.datetime.strptime(datepicker1, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(datepicker2, '%Y-%m-%d').date()

        payment_list = payment_list.filter(created_at__range=[datetime.datetime(date1.year, date1.month, date1.day, 8, 15, 12, 0, pytz.UTC), datetime.datetime(date2.year, date2.month, date2.day, 8, 15, 12, 0, pytz.UTC)])

    payments=[]

    for payment in payment_list[int(start):][:int(length)]:

        payment_details={}
        payment_details['firstname']=payment.user.user.first_name
        payment_details['lastname']=payment.user.user.last_name
        payment_details['username']=payment.user.user.username
        payment_details['email']=payment.user.user.email
        payment_details['date']=(payment.created_at).strftime("%Y/%m/%d %H:%M:%S")
        payment_details['project']=payment.project.title
        if payment.user_reinvestment or payment.admin_reinvestment:
            payment_details['amount']=0
        else:
            payment_details['amount']=payment.amount
        if payment.user_reinvestment:
            payment_details['user_reinvestment'] = round(payment.user_reinvestment.amount, 2)
        else:
            payment_details['user_reinvestment'] = 0
        if payment.admin_reinvestment:
            payment_details['admin_reinvestment'] = round(payment.amount, 2)
        else:
            payment_details['admin_reinvestment']=0
        if payment.tip:
            payment_details['tip'] = round(payment.tip.amount, 2)
        else:
            payment_details['tip'] = 0
        if payment.tip and payment.amount:
            payment_details['total']=round(payment.tip.amount+payment.amount, 2)
        if payment.tip and not payment.amount:
            payment_details['total']=round(payment.tip.amount, 2)
        if payment.amount and not payment.tip:
            payment_details['total']=round(payment.amount, 2)
        if not payment.amount and not payment.tip:
            payment_details['total'] = 0
        payments.append(payment_details)

    json_response={ "draw": draw, "recordsTotal": Payment.objects.filter(project=project).count(), "recordsFiltered": payment_list.count(), "data": payments }

    return HttpResponse(json.dumps(json_response), content_type='application/json')


def payment_data_table(request):
    draw=request.GET.get('draw')
    datepicker1=request.GET.get('datepicker1')
    datepicker2=request.GET.get('datepicker2')
    length=request.GET.get('length')
    order=request.GET.get('order[0][dir]')
    start=request.GET.get('start')
    search=request.GET.get('search[value]')
    currentSortByCol = request.GET.get('order[0][column]')

    fields=['user__user__first_name', 'user__user__last_name', 'user__user__username','user__user__email','created_at','project__title','amount','user_reinvestment__amount','admin_reinvestment__amount','tip__amount']

    order_by={'desc':'-','asc':''}

    column_order=order_by[order]+fields[int(currentSortByCol)]

    payment_list = Payment.objects.all().order_by((column_order))

    if search.strip():
        payment_list = payment_list.filter(Q(user__user__username__icontains=search)|
                                              Q(user__user__first_name__icontains=search)|
                                              Q(project__title__icontains=search)|
                                              Q(amount__icontains=search)|
                                              Q(user__user__last_name__icontains=search)|
                                              Q(user__user__email__icontains=search))


    if datepicker1 and datepicker2:
        import datetime
        import pytz

        date1 = datetime.datetime.strptime(datepicker1, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(datepicker2, '%Y-%m-%d').date()

        payment_list = payment_list.filter(created_at__range=[datetime.datetime(date1.year, date1.month, date1.day, 8, 15, 12, 0, pytz.UTC), datetime.datetime(date2.year, date2.month, date2.day+1, 8, 15, 12, 0, pytz.UTC)])

    payments=[]

    for payment in payment_list[int(start):][:int(length)]:

        payment_details={}
        payment_details['firstname']=payment.user.user.first_name
        payment_details['lastname']=payment.user.user.last_name
        payment_details['username']=payment.user.user.username
        payment_details['email']=payment.user.user.email
        payment_details['date']=(payment.created_at).strftime("%Y/%m/%d %H:%M:%S")
        payment_details['project']=payment.project.title
        if payment.user_reinvestment or payment.admin_reinvestment:
            payment_details['amount']=0
        else:
            payment_details['amount']=payment.amount
        if payment.user_reinvestment:
            payment_details['user_reinvestment'] = round(payment.user_reinvestment.amount, 2)
        else:
            payment_details['user_reinvestment'] = 0
        if payment.admin_reinvestment:
            payment_details['admin_reinvestment'] = round(payment.amount, 2)
        else:
            payment_details['admin_reinvestment']=0
        if payment.tip:
            payment_details['tip'] = round(payment.tip.amount, 2)
        else:
            payment_details['tip'] = 0
        if payment.tip and payment.amount:
            payment_details['total']=round(payment.tip.amount+payment.amount, 2)
        if payment.tip and not payment.amount:
            payment_details['total']=round(payment.tip.amount, 2)
        if payment.amount and not payment.tip:
            payment_details['total']=round(payment.amount, 2)
        if not payment.amount and not payment.tip:
            payment_details['total'] = 0
        payments.append(payment_details)

    json_response={ "draw": draw, "recordsTotal": Payment.objects.all().count(), "recordsFiltered": payment_list.count(), "data": payments }

    return HttpResponse(json.dumps(json_response), content_type='application/json')


def repayment_table(request):
    draw=request.GET.get('draw')
    datepicker1=request.GET.get('datepicker1')
    datepicker2=request.GET.get('datepicker2')
    length=request.GET.get('length')
    order=request.GET.get('order[0][dir]')
    start=request.GET.get('start')
    search=request.GET.get('search[value]')
    currentSortByCol = request.GET.get('order[0][column]')

    fields=['user__user__first_name', 'user__user__last_name', 'user__user__username','user__user__email','created_at','amount','admin_repayment__amount','project__title']

    order_by={'desc':'-','asc':''}

    column_order=order_by[order]+fields[int(currentSortByCol)]

    repayment_list = RepaymentFragment.objects.filter(amount__gt=0.00).order_by((column_order))

    if search.strip():
        repayment_list = repayment_list.filter(Q(user__user__username__icontains=search)|
                                              Q(user__user__first_name__icontains=search)|
                                              Q(project__title__icontains=search)|
                                              Q(amount__icontains=search)|
                                              Q(user__user__last_name__icontains=search)|
                                              Q(user__user__email__icontains=search))


    if datepicker1 and datepicker2:
        import datetime
        import pytz

        date1 = datetime.datetime.strptime(datepicker1, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(datepicker2, '%Y-%m-%d').date()

        repayment_list = repayment_list.filter(created_at__range=[datetime.datetime(date1.year, date1.month, date1.day, 8, 15, 12, 0, pytz.UTC), datetime.datetime(date2.year, date2.month, date2.day, 8, 15, 12, 0, pytz.UTC)])

    payments=[]

    for payment in repayment_list[int(start):][:int(length)]:
        payment_details={}
        payment_details['firstname']=payment.user.user.first_name
        payment_details['lastname']=payment.user.user.last_name
        payment_details['username']=payment.user.user.username
        payment_details['email']=payment.user.user.email
        payment_details['date']=(payment.created_at).strftime("%Y/%m/%d %H:%M:%S")
        payment_details['project']=payment.project.title
        payment_details['donated_amount']=round(Payment.objects.filter(user=payment.user).filter(project=payment.project).aggregate(Sum('amount'))['amount__sum'] or 0,2)
        payment_details['repayment_amount']=round(payment.amount,2)

        payments.append(payment_details)

    json_response={ "draw": draw, "recordsTotal": RepaymentFragment.objects.all().count(), "recordsFiltered": repayment_list.count(), "data": payments }

    return HttpResponse(json.dumps(json_response), content_type='application/json')


def export_csv(request):
    """
    Export financial report CSV from the admin side.
    :param request:
    :return: Generate CSV report of selected records.
    """
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=RE-volv_report.csv'
    payments = Payment.objects.all().order_by('-created_at')
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"FIRST NAME"),
        smart_str(u"LAST NAME"),
        smart_str(u"USERNAME"),
        smart_str(u"EMAIL"),
        smart_str(u"DATE"),
        smart_str(u"NAME OF PROJECT"),
        smart_str(u"DONATION TO SOLAR SEED FUND"),
        smart_str(u"REINVESTMENT IN SOLAR SEED FUND"),
        smart_str(u"ADMIN REINVESTMENT IN SOLAR SEED FUND"),
        smart_str(u"DONATION TO OPERATION"),
        smart_str(u"TOTAL DONATIONS"),

    ])

    for payment in payments:
        if payment.admin_reinvestment:
            admin_reinvestment=round(payment.amount,2)
        else:
            admin_reinvestment=0

        if payment.user_reinvestment:
            user_reinvestment=round(payment.user_reinvestment.amount, 2)
        else:
            user_reinvestment=0

        if payment.admin_reinvestment or payment.user_reinvestment:
            donation_amount=0
        else:
            donation_amount = payment.amount

        if payment.tip:
            tip=round(payment.tip.amount,2)
        else:
            tip=0

        if payment.tip and payment.amount:
            total = round(payment.tip.amount + payment.amount, 2)
        if payment.tip and not payment.amount:
            total = round(payment.tip.amount, 2)
        if payment.amount and not payment.tip:
            total = round(payment.amount, 2)
        if not payment.amount and not payment.tip:
            total = 0

        writer.writerow([
            smart_str(payment.user.user.first_name),
            smart_str(payment.user.user.last_name),
            smart_str(payment.user.user.username),
            smart_str(payment.user.user.email),
            smart_str(payment.created_at),
            smart_str(payment.project.title),
            smart_str(donation_amount),
            smart_str(user_reinvestment),
            smart_str(admin_reinvestment),
            smart_str(tip),
            smart_str(total),
        ])
    return response

def export_xlsx(request):
    """
       Export financial report Excel from the admin side.
       :param request:
       :return: Generate Excel report of selected records.
       """
    import openpyxl
    try:
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter

    payments = Payment.objects.all().order_by('-created_at')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=RE-volv.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "RE-volv"

    row_num = 0

    columns = [
        (u"FIRST NAME",30),
        (u"LAST NAME",30),
        (u"USERNAME",30),
        (u"EMAIL",30),
        (u"DATE",30),
        (u"NAME OF PROJECT",30),
        (u"DONATION TO SOLAR SEED FUND",30),
        (u"REINVESTMENT IN SOLAR SEED FUND",20),
        (u"ADMIN REINVESTMENT IN SOLAR SEED FUND",20),
        (u"DONATION TO OPERATION",20),
        (u"TOTAL DONATIONS",20),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for payment in payments:
        if payment.admin_reinvestment:
            admin_reinvestment = round(payment.amount, 2)
        else:
            admin_reinvestment = 0

        if payment.user_reinvestment:
            user_reinvestment = round(payment.user_reinvestment.amount, 2)
        else:
            user_reinvestment = 0

        if payment.admin_reinvestment or payment.user_reinvestment:
            donation_amount = 0
        else:
            donation_amount = payment.amount

        if payment.tip:
            tip = round(payment.tip.amount, 2)
        else:
            tip = 0

        if payment.tip and payment.amount:
            total = round(payment.tip.amount + payment.amount, 2)
        if payment.tip and not payment.amount:
            total = round(payment.tip.amount, 2)
        if payment.amount and not payment.tip:
            total = round(payment.amount, 2)
        if not payment.amount and not payment.tip:
            total = 0

        row_num += 1
        row = [
            payment.user.user.first_name,
            payment.user.user.last_name,
            payment.user.user.username,
            payment.user.user.email,
            payment.created_at,
            payment.project.title,
            donation_amount,
            user_reinvestment,
            admin_reinvestment,
            tip,
            total,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]

    wb.save(response)
    return response

def export_repayment_csv(request):
    """
    Export financial report CSV from the admin side.
    :param request:
    :return: Generate CSV report of selected records.
    """
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Repayment_report.csv'
    repayments = RepaymentFragment.objects.filter(amount__gt=0.00).order_by('-created_at')
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"FIRST NAME"),
        smart_str(u"LAST NAME"),
        smart_str(u"USERNAME"),
        smart_str(u"EMAIL"),
        smart_str(u"DATE"),
        smart_str(u"NAME OF PROJECT"),
        smart_str(u"DONATION AMOUNT"),
        smart_str(u"REPAYMENT AMOUNT"),


    ])

    for payment in repayments:

        writer.writerow([
            smart_str(payment.user.user.first_name),
            smart_str(payment.user.user.last_name),
            smart_str(payment.user.user.username),
            smart_str(payment.user.user.email),
            smart_str(payment.created_at),
            smart_str(payment.project.title),
            smart_str(round(Payment.objects.filter(user=payment.user).filter(project=payment.project).aggregate(Sum('amount'))['amount__sum'] or 0,2)),
            smart_str(round(payment.amount,2)),
        ])
    return response

def export_repayment_xlsx(request):
    """
       Export financial report Excel from the admin side.
       :param request:
       :return: Generate Excel report of selected records.
       """
    import openpyxl
    try:
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter

    repayments = RepaymentFragment.objects.filter(amount__gt=0.00).order_by('-created_at')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Repayment_report.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "RE-volv"

    row_num = 0

    columns = [
        (u"FIRST NAME",30),
        (u"LAST NAME",30),
        (u"USERNAME",30),
        (u"EMAIL",30),
        (u"DATE",30),
        (u"NAME OF PROJECT",30),
        (u"DONATION AMOUNT",30),
        (u"REPAYMENT AMOUNT",30),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for payment in repayments:
        row_num += 1
        row = [
            payment.user.user.first_name,
            payment.user.user.last_name,
            payment.user.user.username,
            payment.user.user.email,
            payment.created_at,
            payment.project.title,
            round(Payment.objects.filter(user=payment.user).filter(project=payment.project).aggregate(Sum('amount'))['amount__sum'] or 0, 2),
            round(payment.amount, 2)

        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]

    wb.save(response)
    return response

def add_email_to_mailing_list(request):
    if request.POST['email']:
        is_email_exist = False
        email_address = request.POST['email']
        list = mailchimp.utils.get_connection().get_list_by_id(LIST_ID)
        for resp in list.con.list_members(list.id)['data']:
            if email_address == resp['email']:
                is_email_exist = True
        if is_email_exist:
            return HttpResponse(json.dumps({'status': 'already_exist'}), content_type="application/json")
        else:
            try:
                list.subscribe(email_address, {'EMAIL': email_address})
            except Exception:
                return HttpResponse(json.dumps({'status': 'subscription_fail'}), content_type="application/json")
            return HttpResponse(json.dumps({'status': 'subscription_success'}), content_type="application/json")
    else:
        return HttpResponse(json.dumps({'status': 'subscription_fail'}), content_type="application/json")




