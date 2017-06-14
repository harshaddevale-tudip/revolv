from django.contrib import admin
from django.http import HttpResponse

from .models import (
    AdminReinvestment, AdminRepayment, Payment, ProjectMontlyRepaymentConfig,
    PaymentType, RepaymentFragment, UserReinvestment, Tip
)

admin.site.register(AdminReinvestment)
admin.site.register(AdminRepayment)
admin.site.register(Payment)
admin.site.register(ProjectMontlyRepaymentConfig)
admin.site.register(PaymentType)
admin.site.register(RepaymentFragment)
admin.site.register(UserReinvestment)
admin.site.register(Tip)

def export_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=RE-volv_report.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"FIRST NAME"),
        smart_str(u"LAST NAME"),
        smart_str(u"EMAIL"),
        smart_str(u"DATE"),
        smart_str(u"NAME OF PROJECT"),
        smart_str(u"DONATION TO SOLAR SEED FUND"),
        smart_str(u"REINVESTMENT IN SOLAR SEED FUND"),
        smart_str(u"ADMIN REINVESTMENT IN SOLAR SEED FUND"),
        smart_str(u"DONATION TO OPERATION"),
        smart_str(u"TOTAL DONATIONS"),

    ])

    for payment in queryset:
        if payment.admin_reinvestment:
            admin_reinvestment=round(payment.amount,2)
        else:
            admin_reinvestment=0

        if payment.user_reinvestment:
            user_reinvestment=round(payment.user_reinvestment.amount, 2)
        else:
            user_reinvestment=0

        if payment.admin_reinvestment or payment.user_reinvestment:
            donation_amount=0
        else:
            donation_amount = payment.amount

        if payment.tip:
            tip=round(payment.tip.amount,2)
        else:
            tip=0

        if payment.tip and payment.amount:
            total = round(payment.tip.amount + payment.amount, 2)
        if payment.tip and not payment.amount:
            total = round(payment.tip.amount, 2)
        if payment.amount and not payment.tip:
            total = round(payment.amount, 2)
        if not payment.amount and not payment.tip:
            total = 0

        writer.writerow([
            smart_str(payment.user.user.first_name),
            smart_str(payment.user.user.last_name),
            smart_str(payment.user.user.email),
            smart_str(payment.created_at),
            smart_str(payment.project.title),
            smart_str(donation_amount),
            smart_str(user_reinvestment),
            smart_str(admin_reinvestment),
            smart_str(tip),
            smart_str(total),
        ])
    return response
export_csv.short_description = u"Export CSV"


def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    try:
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=RE-volv.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "RE-volv"

    row_num = 0

    columns = [
        (u"FIRST NAME",30),
        (u"LAST NAME",30),
        (u"EMAIL",30),
        (u"DATE",30),
        (u"NAME OF PROJECT",30),
        (u"DONATION TO SOLAR SEED FUND",30),
        (u"REINVESTMENT IN SOLAR SEED FUND",20),
        (u"ADMIN REINVESTMENT IN SOLAR SEED FUND",20),
        (u"DONATION TO OPERATION",20),
        (u"TOTAL DONATIONS",20),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for payment in queryset:
        if payment.admin_reinvestment:
            admin_reinvestment = round(payment.amount, 2)
        else:
            admin_reinvestment = 0

        if payment.user_reinvestment:
            user_reinvestment = round(payment.user_reinvestment.amount, 2)
        else:
            user_reinvestment = 0

        if payment.admin_reinvestment or payment.user_reinvestment:
            donation_amount = 0
        else:
            donation_amount = payment.amount

        if payment.tip:
            tip = round(payment.tip.amount, 2)
        else:
            tip = 0

        if payment.tip and payment.amount:
            total = round(payment.tip.amount + payment.amount, 2)
        if payment.tip and not payment.amount:
            total = round(payment.tip.amount, 2)
        if payment.amount and not payment.tip:
            total = round(payment.amount, 2)
        if not payment.amount and not payment.tip:
            total = 0

        row_num += 1
        row = [
            payment.user.user.first_name,
            payment.user.user.last_name,
            payment.user.user.email,
            payment.created_at,
            payment.project.title,
            donation_amount,
            user_reinvestment,
            admin_reinvestment,
            tip,
            total,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]

    wb.save(response)
    return response

export_xlsx.short_description = u"Export Excel"

class Paymentadmin(admin.ModelAdmin):

    search_fields = ('user__user__username','user__user__first_name','user__user__last_name','user__user__email','amount')

    actions = [export_csv, export_xlsx]

admin.site.unregister(Payment)
admin.site.register(Payment, Paymentadmin)






